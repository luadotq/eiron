# EIRON - Extensible Intelligent Retrieval Open-source Navigator
# Testing version
# GNU GPL V3 License
# Created by - Koger
import os
import re
import mmh3
import time
import pickle
import zlib
import heapq
import threading
import psutil
import resource
from dataclasses import dataclass
from typing import Dict, List, Tuple, Generator
from concurrent.futures import ThreadPoolExecutor, as_completed
from queue import Queue, Empty
from pdfminer.high_level import extract_text
from docx import Document
from openpyxl import load_workbook
from html2text import html2text
import pandas as pd
import csv

# Configuration
MAX_RAM_BYTES = 12 * 1024 * 1024 * 1024  # 12 GB Ram
resource.setrlimit(resource.RLIMIT_AS, (MAX_RAM_BYTES, MAX_RAM_BYTES))
CHUNK_SIZE = 250_000
MAX_DOCS_IN_MEMORY = 1500
MIN_TOKEN_LEN = 3
INDEX_FILE = "search_index.pkl"
TMP_INDEX_FILE = "search_index_tmp.pkl"


@dataclass
class DocumentInfo:
    path: str
    excerpt: str
    size: int
    last_modified: float
    chunk_id: int = 0


class HybridSearchEngine:
    def __init__(self):
        self.inverted_index: Dict[str, Dict[Tuple[int, int], int]] = {}
        self.documents: Dict[int, DocumentInfo] = {}
        self.doc_path_to_id: Dict[str, int] = {}
        self.next_doc_id = 1
        self.index_queue = Queue(maxsize=5000)
        self.shutdown_flag = False
        self.lock = threading.Lock()
        self.last_flush = time.time()

        # Index load
        self._load_index()

        # Threading func
        self._start_workers()

    def _start_workers(self):
#Threading
        num_workers = min(8, (os.cpu_count() or 4) + 2)
        self.workers = []
        for _ in range(num_workers):
            t = threading.Thread(target=self._process_queue, daemon=True)
            t.start()
            self.workers.append(t)

        self.memory_manager = threading.Thread(target=self._manage_memory, daemon=True)
        self.memory_manager.start()

    def _manage_memory(self):
#Memory management
        while not self.shutdown_flag:
            try:
                mem = psutil.virtual_memory()
                used_percent = mem.used / MAX_RAM_BYTES

                # Dynamic threshold
                threshold = 0.75 + (0.15 * (len(self.documents) / MAX_DOCS_IN_MEMORY))

                if used_percent > threshold or \
                        (len(self.documents) >= MAX_DOCS_IN_MEMORY and time.time() - self.last_flush > 60):
                    self._flush_to_disk()
                    self.last_flush = time.time()

                # Adaptive sleep
                sleep_time = 0.2 if used_percent > 0.7 else 1.0
                time.sleep(sleep_time)

            except Exception as e:
                print(f"[-] Memory manager error: {e}")
                break

    def _flush_to_disk(self):
#Adaptive flush
        with self.lock:
            if not self.documents:
                return

            print(f"Drop {len(self.documents)} saves to storage")
            start_time = time.time()

            try:
                # Serializing feature
                data = {
                    'inverted_index': self.inverted_index,
                    'documents': self.documents,
                    'doc_path_to_id': self.doc_path_to_id,
                    'next_doc_id': self.next_doc_id
                }

                # Compression and atomic writes
                with open(TMP_INDEX_FILE, 'wb') as f:
                    f.write(zlib.compress(pickle.dumps(data, protocol=pickle.HIGHEST_PROTOCOL)))

                os.replace(TMP_INDEX_FILE, INDEX_FILE)

                # Memory clearing feature
                self.inverted_index.clear()
                self.documents.clear()
                self.doc_path_to_id.clear()

                print(f"Clear completed in {time.time() - start_time:.2f} s")

            except Exception as e:
                print(f"[-] Clear Failure {e}")
                # Emergency memory drop
                self.inverted_index = {}
                self.documents = {}
                self.doc_path_to_id = {}

    def _load_index(self):
#Index load feature
        if os.path.exists(INDEX_FILE):
            try:
                print("Loading Index...")
                with open(INDEX_FILE, 'rb') as f:
                    data = pickle.loads(zlib.decompress(f.read()))
                    self.inverted_index = data['inverted_index']
                    self.documents = data['documents']
                    self.doc_path_to_id = data['doc_path_to_id']
                    self.next_doc_id = data['next_doc_id']
                print(f"Index Loaded: {len(self.documents)} Saves")
            except Exception as e:
                print(f"[-] Failed to load index {e}")

    def _process_queue(self):
#queue processing
        while not self.shutdown_flag:
            try:
                doc_info, tokens = self.index_queue.get(timeout=1)

                with self.lock:
                    doc_id = self.doc_path_to_id.get(doc_info.path, self.next_doc_id)
                    if doc_id == self.next_doc_id:
                        self.doc_path_to_id[doc_info.path] = doc_id
                        self.next_doc_id += 1

                    self.documents[doc_id] = doc_info

                    for token, count in tokens.items():
                        if token not in self.inverted_index:
                            self.inverted_index[token] = {}
                        key = (doc_id, doc_info.chunk_id)
                        self.inverted_index[token][key] = count

            except Empty:
                continue
            except Exception as e:
                print(f"[-] Queue error: {e}")

    def _tokenize(self, text: str) -> Dict[str, int]:
#Tokenize + Simplify feature
        # Fixing Context
        text = re.sub(r'\s+', ' ', text.lower())
        # Tokenize
        tokens = re.findall(r'\b[\w\'-]{%d,}\b' % MIN_TOKEN_LEN, text)

        counts = {}
        for t in tokens:
            if 3 <= len(t) <= 25:
                counts[t] = counts.get(t, 0) + 1
        return counts

    def _parse_file(self, filepath: str) -> Generator[Tuple[str, int], None, None]:
#Advanced parsing feature
        try:
            ext = os.path.splitext(filepath)[1].lower()
            file_size = os.path.getsize(filepath)

            if ext == '.pdf':
                text = extract_text(filepath)
                for i in range(0, len(text), CHUNK_SIZE):
                    yield text[i:i + CHUNK_SIZE], i // CHUNK_SIZE

            elif ext == '.docx':
                doc = Document(filepath)
                full_text = '\n'.join(p.text for p in doc.paragraphs if p.text)
                for i in range(0, len(full_text), CHUNK_SIZE):
                    yield full_text[i:i + CHUNK_SIZE], i // CHUNK_SIZE

            elif ext in ('.xls', '.xlsx'):
                if ext == '.xlsx':
                    wb = load_workbook(filepath, read_only=True)
                    for sheet in wb.sheetnames:
                        rows = []
                        for row in wb[sheet].iter_rows(values_only=True):
                            rows.append(' '.join(str(c) for c in row if c))
                        text = ' '.join(rows)
                        for i in range(0, len(text), CHUNK_SIZE):
                            yield text[i:i + CHUNK_SIZE], i // CHUNK_SIZE
                else:
                    df = pd.read_excel(filepath, sheet_name=None)
                    for sheet in df.values():
                        text = sheet.to_string()
                        for i in range(0, len(text), CHUNK_SIZE):
                            yield text[i:i + CHUNK_SIZE], i // CHUNK_SIZE

            elif ext == '.html':
                with open(filepath, 'r', encoding='utf-8') as f:
                    text = html2text(f.read())
                    for i in range(0, len(text), CHUNK_SIZE):
                        yield text[i:i + CHUNK_SIZE], i // CHUNK_SIZE

            elif ext == '.csv':
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    chunk = []
                    size = 0
                    chunk_id = 0
                    for row in reader:
                        row_text = ' '.join(str(c) for c in row)
                        if size + len(row_text) > CHUNK_SIZE and chunk:
                            yield ' '.join(chunk), chunk_id
                            chunk = []
                            size = 0
                            chunk_id += 1
                        chunk.append(row_text)
                        size += len(row_text)
                    if chunk:
                        yield ' '.join(chunk), chunk_id

            else:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    chunk_id = 0
                    while True:
                        chunk = f.read(CHUNK_SIZE)
                        if not chunk:
                            break
                        yield chunk, chunk_id
                        chunk_id += 1

        except Exception as e:
            print(f"[-] Failed to process {filepath}: {e}")

    def index_file(self, filepath: str):
#Advanced indexing algorithm
        try:
            stat = os.stat(filepath)
            file_size = stat.st_size
            excerpt = ""
            processed_chunks = 0

            for chunk, chunk_id in self._parse_file(filepath):
                if not excerpt:
                    excerpt = (chunk[:300] + '...') if len(chunk) > 300 else chunk

                tokens = self._tokenize(chunk)
                if tokens:
                    doc_info = DocumentInfo(
                        path=os.path.abspath(filepath),
                        excerpt=excerpt,
                        size=file_size,
                        last_modified=stat.st_mtime,
                        chunk_id=chunk_id
                    )

                    # 2-Step memory manager
                    mem = psutil.virtual_memory()
                    if mem.used > MAX_RAM_BYTES * 0.85:
                        time.sleep(0.5)

                    self.index_queue.put((doc_info, tokens))
                    processed_chunks += 1

            print(f"File {os.path.basename(filepath)} indexed in ({processed_chunks} chunks)")

        except Exception as e:
            print(f"[-] Indexing error {filepath}: {e}")

    def index_directory(self, directory: str):
#Threaded directory indexing
        if not os.path.isdir(directory):
            print(f"[-] {directory} cannot use for directory")
            return

        print(f"Indexing {directory}...")
        start_time = time.time()
        file_count = 0

        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = []
            for root, _, files in os.walk(directory):
                for file in files:
                    filepath = os.path.join(root, file)
                    futures.append(executor.submit(self.index_file, filepath))
                    file_count += 1

                    # Just progress bar
                    if file_count % 10 == 0:
                        print(f"Processed: {file_count}")

            # Wait for complete
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    print(f"[-] Indexing failure {e}")

        # Final flushing
        self._flush_to_disk()
        print(f"Indexing complete. Files: {file_count}, Time: {time.time() - start_time:.1f} s")

    def search(self, query: str, limit: int = 10) -> List[Dict]:
#Advanced search
        if not query.strip():
            return []

        #Prompt tokenizing
        query_tokens = re.findall(r'\b[\w\'-]{%d,}\b' % MIN_TOKEN_LEN, query.lower())
        if not query_tokens:
            return []

        start_time = time.time()

        # Index loading (if needed)
        if not self.documents:
            self._load_index()

        # Search and ranging
        doc_scores = {}
        with self.lock:
            for token in query_tokens:
                if token in self.inverted_index:
                    for (doc_id, _), count in self.inverted_index[token].items():
                        doc_scores[doc_id] = doc_scores.get(doc_id, 0) + count

        # Scoring feature
        results = []
        for doc_id, score in heapq.nlargest(limit, doc_scores.items(), key=lambda x: x[1]):
            doc = self.documents.get(doc_id)
            if doc:
                # Score factor
                length_factor = min(1.0, 5000 / (doc.size + 1000))
                final_score = score * length_factor

                results.append({
                    'path': doc.path,
                    'excerpt': doc.excerpt,
                    'score': final_score,
                    'modified': time.strftime('%Y-%m-%d %H:%M', time.localtime(doc.last_modified))
                })

        print(f"Find result in {time.time() - start_time:.3f} s. Result: {len(results)}")
        return results

    def shutdown(self):
#Shutown func
        print("Shutdown...")
        self.shutdown_flag = True
        self._flush_to_disk()
        for t in self.workers:
            t.join(timeout=1)
        self.memory_manager.join(timeout=1)


if __name__ == "__main__":
    engine = HybridSearchEngine()
    try:
        while True:
            print("OSSE")
            print("\n1. Reindex\n2. Search\n3. Exit")
            choice = input(" > ").strip()

            if choice == "1":
                path = input("Directory path: ").strip()
                if os.path.isdir(path):
                    engine.index_directory(path)
                else:
                    print("[-] Path doesn't exist")
            elif choice == "2":
                query = input("Prompt: ").strip()
                results = engine.search(query)
                for i, r in enumerate(results, 1):
                    print(f"\n{i}. {r['path']} (relevance: {r['score']:.1f})")
                    print(f"   {r['excerpt']}")
            elif choice == "3":
                break
    finally:
        engine.shutdown()
